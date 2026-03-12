import openai
import openpyxl
import time
import os

# ── CONFIGURATION ──────────────────────────────────────────
API_KEY = os.environ.get("OPENAI_API_KEY", "your-key-here")
EXCEL_FILE = "Prompts_Excel.xlsx"
OUTPUT_FILE = "Results.xlsx"

# ── LOAD POLICY DOCUMENTS ──────────────────────────────────
def load_policies():
    policies = {}
    files = {
        "cd_policy": "cd_policy.txt",
        "dispute_policy": "account_dispute_policy.txt",
        "business_policy": "business_account_policy.txt"
    }
    for name, filename in files.items():
        with open(filename, "r") as f:
            policies[name] = f.read()
    return policies

# ── BUILD SYSTEM PROMPT ────────────────────────────────────
def build_system_prompt(policies):
    return f"""You are a customer service AI assistant for FirstBank.
You can only answer questions based on the following policy documents.
If the answer is not in these documents, say so clearly.
Do not make up information.

=== CD POLICY ===
{policies['cd_policy']}

=== DISPUTE POLICY ===
{policies['dispute_policy']}

=== BUSINESS ACCOUNT POLICY ===
{policies['business_policy']}
"""

# ── ASK ONE QUESTION IN A FRESH CONVERSATION ───────────────
def ask_question(client, system_prompt, question):
    response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        max_tokens=500,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": question}
        ]
    )
    return response.choices[0].message.content

# ── MAIN SCRIPT ────────────────────────────────────────────
def main():

    print("Loading policy documents...")
    policies = load_policies()
    system_prompt = build_system_prompt(policies)

    print("Connecting to OpenAI API...")
    client = openai.OpenAI(api_key=API_KEY)

    print("Loading prompts from Excel...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Find or create the API Response column
    header_row = 2
    last_col = ws.max_column + 1
    ws.cell(row=header_row, column=last_col, value="API Response")

    # Process each prompt starting from row 3
    total = 0
    for row in ws.iter_rows(min_row=3, values_only=False):
        prompt_id = row[0].value
        prompt_text = row[2].value

        # Skip empty rows
        if not prompt_text:
            continue

        print(f"Running {prompt_id}: {prompt_text[:50]}...")

        try:
            response = ask_question(client, system_prompt, prompt_text)
            # Write response into the new column
            ws.cell(row=row[0].row, column=last_col, value=response)
            total += 1
            print(f"  ✓ Done")

        except Exception as e:
            ws.cell(row=row[0].row, column=last_col, value=f"ERROR: {str(e)}")
            print(f"  ✗ Error: {e}")

        # Wait 1 second between calls to avoid rate limits
        time.sleep(1)

    # Save results to a new file
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Complete! {total} prompts processed.")
    print(f"📄 Results saved to {OUTPUT_FILE}")

if __name__ == "__main__":
    main()